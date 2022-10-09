import openpyxl as openpyxl
from docxtpl import DocxTemplate
import datetime as dt
import random
from database import crud


def get_int_time(date: dt.datetime = dt.datetime(2022, 9, 26, 0, 0, 0)):
    x = (date - dt.datetime(1970, 1, 1, 0, 0, 0)).total_seconds()
    return int(x)


def generate_code() -> str:
    code_chars = '0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz'
    code = ''
    for i in range(random.randint(10, 15)):
        code += random.choice(code_chars)
    return code


def create_report():
    doc = DocxTemplate("special/Absent.docx")
    replace = {'а': 'a', 'А': 'a', 'б': 'b', 'Б': 'b', 'в': 'c', 'В': 'c'}
    classes = crud.get_classes_list()
    context = {}
    all_students = 0
    all_in_lyceum = 0
    in_lyceum = 0
    out_lyceum = 0
    for class_ in classes:
        students_count = crud.get_class_count(class_[2])
        key = f'{replace[class_[1]]}_{class_[0]}'
        key_all = key + '_all'
        key_absent = key + '_absent'
        key_in = key + '_in'
        key_out = key + '_out'
        all_students += students_count[0]
        all_in_lyceum += students_count[1]
        in_lyceum += students_count[2]
        out_lyceum += students_count[3]
        context[key] = students_count[0] - students_count[1]
        context[key_all] = students_count[0]
        context[key_in] = students_count[2]
        context[key_out] = students_count[3]
        users_absent = crud.get_absent_users_by_class(class_[2])
        context[key_absent] = ', '.join([f'{one[0]} ({one[1]})' for one in users_absent])
    context['all_in_lyceum'] = all_students - out_lyceum
    context['all_students'] = all_students
    context['date'] = dt.datetime.now().strftime("%d.%m.%Y")
    doc.render(context)
    file_name = f'{dt.datetime.now().strftime("%d.%m.%Y")}.docx'
    doc.save(file_name)
    return file_name


def create_task_report():
    wb = openpyxl.load_workbook('special/report.xlsx')
    tasks = crud.get_all_task('all')
    months = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
              'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
    for task in tasks:
        sheet = months[int(task.start_date.strftime("%m")) - 1]
        if sheet not in wb.sheetnames:
            source = wb.get_sheet_by_name('Шаблон')
            target = wb.copy_worksheet(source)
            target.title = sheet
        wb.active = wb.get_sheet_by_name(sheet)
        ws = wb.active
        row = ws.max_row + 1
        ws.cell(column=1, row=row, value=f'{task.start_date.strftime("%d.%m.%Y %H:%M")}')
        user = crud.get_user(user_id=task.client_id)
        ws.cell(column=2, row=row, value=f'{user.get_full_name()}')
        ws.cell(column=3, row=row, value=f'{task.block}')
        ws.cell(column=4, row=row, value=f'{task.cabinet}')
        ws.cell(column=5, row=row, value=f'{task.description}')
        user = crud.get_user(user_id=task.employee)
        ws.cell(column=6, row=row, value=f'{user.get_full_name()}')
        ws.cell(column=7, row=row, value=f'{task.end_date}')
        ws.cell(column=8, row=row, value=f'{task.postponed}')
    wb.remove_sheet(wb.get_sheet_by_name('Шаблон'))
    file_name = f'complete.xlsx'
    wb.save(filename=file_name)
    return file_name

